import time
import logging
import glob
import os
import traceback
from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, ElementClickInterceptedException, StaleElementReferenceException
)

load_dotenv(os.path.join(os.path.dirname(__file__), "config.env"))

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S"
)
log = logging.getLogger(__name__)

NIKSHAY_USERNAME = os.getenv("NIKSHAY_USERNAME", "")
NIKSHAY_PASSWORD = os.getenv("NIKSHAY_PASSWORD", "")
DATA_FOLDER = os.path.dirname(__file__)
HEADLESS = os.getenv("NIKSHAY_HEADLESS", "false").lower() == "true"
START_ROW = int(os.getenv("NIKSHAY_START_ROW", "1"))
DEFAULT_ADDRESS = os.getenv("NIKSHAY_ADDRESS", "Bari Gulni")
DEFAULT_PINCODE = os.getenv("NIKSHAY_PINCODE", "805124")
DEFAULT_HEALTH_FACILITY = os.getenv("NIKSHAY_HEALTH_FACILITY", "Ab HWC Gulni")
WAIT_AFTER_SUBMIT = int(os.getenv("NIKSHAY_WAIT_AFTER_SUBMIT", "15"))

SYMPTOM_LIST = [
    "Asymptomatic", "Cough for more than 2 weeks", "Coughing up blood",
    "Fever", "Weight loss", "Night sweat", "Chest pain",
    "Shortness of breath", "Fatigue",
    "Failure to gain weight (in children)",
    "Decreased activity or playfulness (in children)",
    "Others"
]

SYMPTOM_ALIASES = {
    "cough": "Cough for more than 2 weeks",
    "cough for more than 2 weeks": "Cough for more than 2 weeks",
    "cough for more than two weeks": "Cough for more than 2 weeks",
    "chronic cough": "Cough for more than 2 weeks",
    "persistent cough": "Cough for more than 2 weeks",
    "night sweat": "Night sweat",
    "night sweats": "Night sweat",
    "fever": "Fever",
    "high fever": "Fever",
    "low grade fever": "Fever",
    "chest pain": "Chest pain",
    "chest ache": "Chest pain",
    "weight loss": "Weight loss",
    "weight decrease": "Weight loss",
    "shortness of breath": "Shortness of breath",
    "breathlessness": "Shortness of breath",
    "breathing difficulty": "Shortness of breath",
    "fatigue": "Fatigue",
    "weakness": "Fatigue",
    "tiredness": "Fatigue",
    "blood in sputum": "Coughing up blood",
    "hemoptysis": "Coughing up blood",
    "coughing blood": "Coughing up blood",
    "loss of appetite": "Others",
    "no appetite": "Others",
    "appetite loss": "Others",
    "wound": "Others",
    "ulcer": "Others",
    "swelling": "Others",
    "asymptomatic": "Asymptomatic",
}

KEY_POP_MAP = {
    "Not Applicable": "Not Applicable",
    "Pregnancy": "Pregnancy",
    "pregnancy": "Pregnancy",
    "Elderly (age >60 years)": "Elderly (age >60 years)",
    "elderly": "Elderly (age >60 years)",
    "Elderly": "Elderly (age >60 years)",
    "Other": "Other",
    "other": "Other",
    "Migrant": "Migrant",
    "migrant": "Migrant",
    "Miner": "Miner",
    "miner": "Miner",
    "Prison": "Prison",
    "prison": "Prison",
    "Diabetes": "Diabetes",
    "diabetes": "Diabetes",
    "Cancer": "Cancer",
    "cancer": "Cancer",
    "Hypertensive": "Hypertensive",
    "hypertensive": "Hypertensive",
    "Lactating mother": "Lactating mother",
    "lactating mother": "Lactating mother",
    "Tobacco/smoker": "Tobacco/smoker",
    "tobacco": "Tobacco/smoker",
    "smoker": "Tobacco/smoker",
    "Urban Slum": "Urban Slum",
    "urban slum": "Urban Slum",
    "Contact of Known TB Patients": "Contact of Known TB Patients",
}

MAX_RETRIES = 3


def init_driver():
    options = webdriver.ChromeOptions()
    if HEADLESS:
        options.add_argument("--headless=new")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)


def wait_and_find(driver, by, value, timeout=10):
    return WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, value)))


def wait_clickable(driver, by, value, timeout=10):
    return WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, value)))


def take_screenshot(driver, name):
    try:
        path = os.path.join(DATA_FOLDER, f"debug_{name}.png")
        driver.save_screenshot(path)
        log.info(f"  Screenshot: {path}")
    except Exception as e:
        log.warning(f"  Screenshot failed: {e}")


def debug_page(driver):
    try:
        inputs = driver.find_elements(By.TAG_NAME, "input")
        log.info(f"  Total <input> elements: {len(inputs)}")
        for i, inp in enumerate(inputs):
            try:
                log.info(f"    [{i}] id={inp.get_attribute('id')} name={inp.get_attribute('name')} type={inp.get_attribute('type')} placeholder={inp.get_attribute('placeholder')}")
            except:
                pass
        iframes = driver.find_elements(By.TAG_NAME, "iframe")
        log.info(f"  Total <iframe> elements: {len(iframes)}")
        for i, iframe in enumerate(iframes):
            try:
                log.info(f"    iframe[{i}] id={iframe.get_attribute('id')} src={iframe.get_attribute('src')}")
            except:
                pass
    except Exception as e:
        log.warning(f"  Debug failed: {e}")


def login(driver):
    log.info("Logging in...")
    driver.get("https://www.nikshay.in/Home/Login")
    time.sleep(5)
    take_screenshot(driver, "01_login_page")
    debug_page(driver)

    # Check iframes
    iframes = driver.find_elements(By.TAG_NAME, "iframe")
    if iframes:
        log.info(f"  Found {len(iframes)} iframe(s), trying to switch...")
        for iframe in iframes:
            try:
                driver.switch_to.frame(iframe)
                log.info(f"  Switched to iframe: {iframe.get_attribute('id')}")
                take_screenshot(driver, "01b_iframe")
                # Try finding username inside iframe
                try:
                    driver.find_element(By.ID, "userName")
                    log.info("  Found userName inside iframe!")
                    break
                except:
                    driver.switch_to.default_content()
            except:
                driver.switch_to.default_content()

    # Try multiple selectors for username
    user_selectors = [
        (By.ID, "userName"),
        (By.ID, "username"),
        (By.NAME, "userName"),
        (By.NAME, "username"),
        (By.NAME, "email"),
        (By.XPATH, "//input[@placeholder='Username' or @placeholder='username' or @placeholder='User Name']"),
        (By.XPATH, "//input[@type='text'][1]"),
        (By.CSS_SELECTOR, "input[type='text']"),
        (By.CSS_SELECTOR, "input[type='email']"),
    ]

    user_el = None
    for by, sel in user_selectors:
        try:
            user_el = WebDriverWait(driver, 3).until(EC.presence_of_element_located((by, sel)))
            log.info(f"  Found username field: {sel}")
            break
        except Exception:
            continue

    if not user_el:
        log.error("  Username field NOT found with any selector!")
        take_screenshot(driver, "01_no_username")
        # Save page source for debugging
        with open(os.path.join(DATA_FOLDER, "debug_page_source.html"), "w", encoding="utf-8") as f:
            f.write(driver.page_source)
        log.info("  Page source saved to debug_page_source.html")
        return False

    # Fill username with JS
    try:
        user_el.click()
        time.sleep(0.3)
        user_el.send_keys(NIKSHAY_USERNAME)
        log.info(f"  Username filled via send_keys: {NIKSHAY_USERNAME}")
    except Exception as e:
        log.warning(f"  send_keys failed: {e}, trying JS...")
        js_set_val(driver, user_el, NIKSHAY_USERNAME)
        log.info("  Username filled via JS")

    take_screenshot(driver, "02_username_filled")

    # Try multiple selectors for password
    pass_selectors = [
        (By.ID, "password"),
        (By.ID, "Password"),
        (By.NAME, "password"),
        (By.NAME, "Password"),
        (By.XPATH, "//input[@type='password']"),
        (By.CSS_SELECTOR, "input[type='password']"),
        (By.XPATH, "//input[@placeholder='Password' or @placeholder='password']"),
    ]

    pass_el = None
    for by, sel in pass_selectors:
        try:
            pass_el = WebDriverWait(driver, 3).until(EC.presence_of_element_located((by, sel)))
            log.info(f"  Found password field: {sel}")
            break
        except Exception:
            continue

    if not pass_el:
        log.error("  Password field NOT found!")
        take_screenshot(driver, "02_no_password")
        return False

    try:
        pass_el.click()
        time.sleep(0.3)
        pass_el.send_keys(NIKSHAY_PASSWORD)
        log.info("  Password filled via send_keys")
    except Exception as e:
        log.warning(f"  send_keys failed: {e}, trying JS...")
        js_set_val(driver, pass_el, NIKSHAY_PASSWORD)
        log.info("  Password filled via JS")

    take_screenshot(driver, "03_both_filled")

    # Switch back to main content if in iframe
    try:
        driver.switch_to.default_content()
    except:
        pass

    # Try multiple login button selectors
    login_btn_selectors = [
        (By.ID, "loginButton-nikshay"),
        (By.ID, "loginButton"),
        (By.ID, "login-btn"),
        (By.XPATH, "//button[contains(@id,'login')]"),
        (By.XPATH, "//button[contains(text(),'Login') or contains(text(),'LOGIN') or contains(text(),'Log In')]"),
        (By.XPATH, "//button[contains(text(),'Sign In') or contains(text(),'SIGN IN')]"),
        (By.XPATH, "//input[@type='submit']"),
        (By.XPATH, "//button[@type='submit']"),
        (By.XPATH, "//a[contains(text(),'Login')]"),
        (By.CSS_SELECTOR, "button.btn-primary"),
        (By.CSS_SELECTOR, "button[type='submit']"),
    ]

    clicked = False
    for by, sel in login_btn_selectors:
        try:
            btn = wait_clickable(driver, by, sel, 3)
            js_click(driver, btn)
            clicked = True
            log.info(f"  Login button clicked: {sel}")
            break
        except Exception:
            continue

    if not clicked:
        # Try clicking with JS
        log.info("  Trying JS click on all buttons...")
        try:
            result = driver.execute_script("""
                var btns = document.querySelectorAll('button, input[type="submit"], a.btn');
                for (var i = 0; i < btns.length; i++) {
                    var txt = btns[i].textContent.toLowerCase();
                    if (txt.includes('login') || txt.includes('sign in') || txt.includes('submit')) {
                        btns[i].click();
                        return 'clicked: ' + btns[i].textContent.trim();
                    }
                }
                return 'no button found';
            """)
            log.info(f"  JS click result: {result}")
            if "clicked" in result:
                clicked = True
        except Exception as e:
            log.warning(f"  JS click failed: {e}")

    if not clicked:
        log.error("  Login button NOT found!")
        take_screenshot(driver, "04_no_login_btn")
        return False

    time.sleep(5)
    take_screenshot(driver, "05_after_login_click")

    # Check if login succeeded
    current_url = driver.current_url.lower()
    log.info(f"  Current URL: {driver.current_url}")

    page_src = driver.page_source[:3000].lower()
    if "invalid" in page_src or "incorrect" in page_src or "failed" in page_src:
        log.error("  Login FAILED - Invalid credentials!")
        take_screenshot(driver, "06_login_failed")
        return False

    if "dashboard" in current_url or "newenrollment" in current_url or "home" in current_url:
        log.info("Login successful!")
        take_screenshot(driver, "06_login_success")
        return True

    # Maybe page is still loading
    try:
        WebDriverWait(driver, 10).until(
            lambda d: "dashboard" in d.current_url.lower()
            or "newenrollment" in d.current_url.lower()
            or len(d.find_elements(By.ID, "userName")) == 0
        )
        log.info("Login successful!")
        take_screenshot(driver, "06_login_success")
        return True
    except TimeoutException:
        log.warning("  URL did not change. Taking screenshot for debug...")
        take_screenshot(driver, "06_timeout")
        return True


def dismiss_enrollment_modals(driver):
    """Enrollment page pe duplicate ya koi bhi modal ho toh dismiss karo."""
    result = driver.execute_script("""
        var closed = 0;
        // Close enrollmentDuplicatesBody modal
        var modals = document.querySelectorAll('.modal');
        for (var i = 0; i < modals.length; i++) {
            var m = modals[i];
            var style = window.getComputedStyle(m);
            if (style.display !== 'none' && style.visibility !== 'hidden') {
                // Try close button
                var closeBtns = m.querySelectorAll('button.close, button[data-dismiss="modal"], .modal-footer button');
                for (var j = 0; j < closeBtns.length; j++) {
                    var txt = closeBtns[j].textContent.trim().toLowerCase();
                    if (txt.includes('cancel') || txt.includes('close') || txt.includes('ok') || txt.includes('no') || closeBtns[j].className.includes('close')) {
                        closeBtns[j].click();
                        closed++;
                        break;
                    }
                }
            }
        }
        // Also hide any remaining visible modals via JS
        var modals2 = document.querySelectorAll('.modal.in, .modal[style*="display: block"]');
        for (var k = 0; k < modals2.length; k++) {
            modals2[k].style.display = 'none';
            modals2[k].classList.remove('in');
            closed++;
        }
        // Remove modal backdrop
        var backdrops = document.querySelectorAll('.modal-backdrop');
        for (var b = 0; b < backdrops.length; b++) {
            backdrops[b].remove();
        }
        document.body.classList.remove('modal-open');
        document.body.style.overflow = 'auto';
        return closed;
    """)
    if result:
        log.info(f"  Dismissed {result} enrollment modal(s)")
    return result


def go_to_enrollment(driver):
    log.info("Opening enrollment page...")
    driver.get("https://www.nikshay.in/Dashboard/NewEnrollment")
    WebDriverWait(driver, 8).until(
        lambda d: "enrollment" in d.current_url.lower() or d.find_elements(By.ID, "firstName")
    )
    dismiss_enrollment_modals(driver)


def js_set_val(driver, el, val):
    driver.execute_script("""
        el = arguments[0]; el.value = arguments[1];
        el.dispatchEvent(new Event('input', {bubbles:true}));
        el.dispatchEvent(new Event('change', {bubbles:true}));
    """, el, val)


def js_click(driver, el):
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    time.sleep(0.1)
    try:
        el.click()
    except (ElementClickInterceptedException, StaleElementReferenceException):
        driver.execute_script("arguments[0].click();", el)
    time.sleep(0.1)


def fill_text(driver, label, value):
    xpath = f'//div[contains(@class,"form-group")]//label[contains(@class,"control-label") and contains(text(),"{label}")]/following::input[1] | //div[contains(@class,"form-group")]//label[contains(@class,"control-label") and contains(text(),"{label}")]/following::textarea[1]'
    try:
        el = wait_and_find(driver, By.XPATH, xpath, 5)
        js_set_val(driver, el, value)
        log.info(f"  OK {label}")
        return True
    except Exception as e:
        log.warning(f"  FAIL {label}: {e}")
        return False


def click_radio_or_checkbox(driver, field_label, option_text):
    xpath = f'//div[contains(@class,"form-group")]//label[contains(@class,"control-label") and contains(text(),"{field_label}")]/ancestor::div[contains(@class,"form-group")]//label[contains(text(),"{option_text}")]'
    try:
        el = wait_and_find(driver, By.XPATH, xpath, 4)
        js_click(driver, el)
        log.info(f"  OK {field_label} -> {option_text}")
        return True
    except Exception:
        log.warning(f"  SKIP {field_label} -> {option_text}")
        return False


def fill_select(driver, label, value):
    xpath = f'//div[contains(@class,"form-group")]//label[contains(@class,"control-label") and contains(text(),"{label}")]/following::select[1]'
    try:
        el = wait_and_find(driver, By.XPATH, xpath, 5)
        Select(el).select_by_visible_text(str(value))
        log.info(f"  OK {label} -> {value}")
        return True
    except Exception:
        try:
            el = driver.find_element(By.XPATH, xpath)
            for opt in Select(el).options:
                if value.lower() in opt.text.lower():
                    Select(el).select_by_visible_text(opt.text)
                    log.info(f"  OK {label} -> {opt.text}")
                    return True
        except Exception:
            pass
        log.warning(f"  FAIL {label}: '{value}' not found")
        return False


def fill_vselect(driver, field_label, value):
    xpath = f'//div[contains(@class,"form-group")]//label[contains(@class,"control-label") and contains(text(),"{field_label}")]/ancestor::div[contains(@class,"form-group")]//input[@type="search"]'
    try:
        inp = wait_and_find(driver, By.XPATH, xpath, 5)
        driver.execute_script("arguments[0].scrollIntoView({block:'center'}); arguments[0].removeAttribute('readonly'); arguments[0].disabled = false;", inp)
        driver.execute_script("arguments[0].click();", inp)
        time.sleep(0.3)

        # Step 1: Try FULL exact value first
        inp.clear()
        inp.send_keys(str(value))
        time.sleep(1.5)
        items = driver.find_elements(By.CSS_SELECTOR, "ul.vs__dropdown-menu li")
        for item in items:
            txt = item.text.strip()
            if txt.lower() == value.lower():
                js_click(driver, item)
                log.info(f"  OK {field_label} -> {txt}")
                return True
        inp.send_keys("\ue00c")
        time.sleep(0.2)

        # Step 2: Try partial keywords
        words = value.split()
        for term in words:
            inp.clear()
            inp.send_keys(str(term))
            time.sleep(1.5)
            items = driver.find_elements(By.CSS_SELECTOR, "ul.vs__dropdown-menu li")
            for item in items:
                txt = item.text.strip()
                if txt.lower() == value.lower():
                    js_click(driver, item)
                    log.info(f"  OK {field_label} -> {txt}")
                    return True
            inp.send_keys("\ue00c")
            time.sleep(0.2)

        log.warning(f"  FAIL {field_label}: no valid option for '{value}'")
        return False
    except Exception as e:
        log.warning(f"  FAIL {field_label}: {e}")
        return False


def is_others_checked(driver, field_label):
    try:
        xpath = f'//div[contains(@class,"form-group")]//label[contains(@class,"control-label") and contains(text(),"{field_label}")]/ancestor::div[contains(@class,"form-group")]//label[contains(text(),"Others")]'
        el = driver.find_element(By.XPATH, xpath)
        return "checked" in (el.get_attribute("class") or "") or el.get_attribute("aria-checked") == "true"
    except Exception:
        return False


def handle_symptom(driver, field_label, symptom_text):
    opt = symptom_text

    # Check aliases first
    if symptom_text.lower() in SYMPTOM_ALIASES:
        opt = SYMPTOM_ALIASES[symptom_text.lower()]
    elif symptom_text not in SYMPTOM_LIST:
        # Partial match
        for o in SYMPTOM_LIST:
            if symptom_text.lower() in o.lower() or o.lower() in symptom_text.lower():
                opt = o
                break
        else:
            # Check if symptom contains any keyword from aliases
            for alias_key, alias_val in SYMPTOM_ALIASES.items():
                if alias_key in symptom_text.lower() or symptom_text.lower() in alias_key:
                    opt = alias_val
                    break
            else:
                opt = "Others"

    if click_radio_or_checkbox(driver, field_label, opt):
        log.info(f"  OK {field_label} -> {opt}")
        if opt == "Others":
            try:
                inp = driver.find_element(By.ID, "othersymptomText")
                js_set_val(driver, inp, symptom_text)
            except Exception:
                pass
        return True
    else:
        if not is_others_checked(driver, field_label):
            if click_radio_or_checkbox(driver, field_label, "Others"):
                log.info(f"  OK {field_label} -> Others (for '{symptom_text}')")
        try:
            inp = driver.find_element(By.ID, "othersymptomText")
            existing = inp.get_attribute("value") or ""
            new_val = (existing + ", " + symptom_text).strip(", ")
            js_set_val(driver, inp, new_val)
        except Exception:
            pass
        return False


def wait_for_page_load(driver, timeout=10):
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
    except TimeoutException:
        pass


def submit_form(driver):
    log.info("  Trying to submit form...")
    wait_for_page_load(driver)

    # JS fallback pehle try karo - fast aur reliable
    try:
        result = driver.execute_script("""
            var btns = document.querySelectorAll('button, input[type="submit"], a.btn');
            for (var i = 0; i < btns.length; i++) {
                var rect = btns[i].getBoundingClientRect();
                if (rect.width === 0 || rect.height === 0) continue;
                var txt = btns[i].textContent.toLowerCase();
                if (txt.includes('add case') && txt.includes('proceed')) {
                    btns[i].click();
                    return 'clicked: ' + btns[i].textContent.trim();
                }
            }
            for (var i = 0; i < btns.length; i++) {
                var rect = btns[i].getBoundingClientRect();
                if (rect.width === 0 || rect.height === 0) continue;
                var txt = btns[i].textContent.toLowerCase();
                if (txt.includes('add case') || txt.includes('proceed to generate')) {
                    btns[i].click();
                    return 'clicked: ' + btns[i].textContent.trim();
                }
            }
            for (var i = 0; i < btns.length; i++) {
                var rect = btns[i].getBoundingClientRect();
                if (rect.width === 0 || rect.height === 0) continue;
                var txt = btns[i].textContent.toLowerCase();
                if (txt.includes('save') || txt.includes('submit')) {
                    btns[i].click();
                    return 'clicked: ' + btns[i].textContent.trim();
                }
            }
            return 'no button found';
        """)
        if "clicked" in result:
            log.info(f"  OK Form submitted: {result}")
            time.sleep(1)
            wait_for_page_load(driver)
            return True
    except Exception as e:
        log.warning(f"  JS submit failed: {e}")

    # XPath fallback
    xpaths = [
        "//button[contains(text(),'ADD CASE') and contains(text(),'ABHA')]",
        "//button[contains(text(),'ADD CASE') and contains(text(),'PROCEED')]",
        "//button[contains(text(),'ADD CASE')]",
        "//button[normalize-space()='Save']",
        "//button[normalize-space()='Submit']",
        "//input[@type='submit']",
    ]
    for btn_xpath in xpaths:
        try:
            btn = wait_clickable(driver, By.XPATH, btn_xpath, 2)
            js_click(driver, btn)
            time.sleep(1)
            wait_for_page_load(driver)
            log.info("  OK Form submitted!")
            return True
        except (TimeoutException, NoSuchElementException):
            continue

    log.warning("  FAIL Submit button not found!")
    return False


def click_add_case(driver):
    btns = [
        "//a[contains(text(),'Add Case')]",
        "//a[contains(text(),'Add New')]",
        "//button[contains(text(),'Add Case')]",
        "//a[contains(text(),'New Enrollment')]",
        "//a[contains(@href,'NewEnrollment')]",
        "//button[contains(@class,'btn') and contains(text(),'Add')]",
    ]
    for attempt in range(20):
        time.sleep(0.5)
        for btn_xpath in btns:
            try:
                btn = driver.find_element(By.XPATH, btn_xpath)
                if btn.is_displayed():
                    js_click(driver, btn)
                    log.info("  OK Clicked Add Case")
                    WebDriverWait(driver, 5).until(
                        lambda d: "enrollment" in d.current_url.lower() or d.find_elements(By.ID, "firstName")
                    )
                    time.sleep(1)
                    return True
            except (NoSuchElementException, StaleElementReferenceException):
                continue
    log.warning("  FAIL Add Case button not found!")
    return False


def dismiss_popups(driver):
    popup_xpaths = [
        "//div[contains(@class,'modal')]//button[contains(text(),'Cancel')]",
        "//div[contains(@class,'modal')]//button[text()='Cancel']",
        "//button[contains(text(),'Cancel')]",
        "//div[contains(@class,'modal')]//button[@data-dismiss='modal']",
        "//div[contains(@class,'modal')]//button[contains(@class,'close')]",
        "//button[contains(text(),'OK')]",
        "//button[contains(text(),'Close')]",
    ]
    dismissed = 0
    for _ in range(8):
        time.sleep(0.5)
        for xpath in popup_xpaths:
            try:
                btn = driver.find_element(By.XPATH, xpath)
                if btn.is_displayed():
                    js_click(driver, btn)
                    dismissed += 1
                    time.sleep(1)
            except Exception:
                continue
    if dismissed:
        log.info(f"  Dismissed {dismissed} popup(s)")


def cancel_abha_popup(driver):
    """After submit, ABHA popup aata hai - usme Cancel button dabao."""
    log.info("  Looking for ABHA popup to cancel...")

    # JS pehle try karo - sabse fast
    try:
        result = driver.execute_script("""
            var btns = document.querySelectorAll('button, a, input[type="button"]');
            for (var i = 0; i < btns.length; i++) {
                var el = btns[i];
                var rect = el.getBoundingClientRect();
                if (rect.width > 0 && rect.height > 0) {
                    var txt = (el.textContent || el.value || '').trim().toLowerCase();
                    if (txt === 'cancel' || txt.includes('cancel')) {
                        el.click();
                        return 'clicked: ' + (el.textContent || el.value).trim();
                    }
                }
            }
            return 'no cancel button found';
        """)
        if "clicked" in result:
            log.info(f"  OK ABHA popup cancelled (JS): {result}")
            time.sleep(0.5)
            return True
    except Exception:
        pass

    # XPath fallback
    cancel_xpaths = [
        "//div[contains(@class,'modal')]//button[contains(text(),'Cancel')]",
        "//button[contains(text(),'Cancel')]",
        "//button[normalize-space()='Cancel']",
    ]
    for _ in range(3):
        time.sleep(0.5)
        for xpath in cancel_xpaths:
            try:
                btn = driver.find_element(By.XPATH, xpath)
                if btn.is_displayed():
                    js_click(driver, btn)
                    log.info("  OK ABHA popup cancelled")
                    time.sleep(0.5)
                    return True
            except Exception:
                continue

    log.warning("  No ABHA popup found (may not have appeared)")
    return False


def click_key_population(driver, value):
    """Key Population/Risk factors ke disabled checkboxes ko JS se enable aur click karo."""
    log.info(f"  Setting Key Population -> {value}")

    # Scroll to Key Population section first
    driver.execute_script("""
        var labels = document.querySelectorAll('label.control-label');
        for (var i = 0; i < labels.length; i++) {
            if (labels[i].textContent.includes('Key Population')) {
                labels[i].scrollIntoView({block: 'center'});
                break;
            }
        }
    """)
    time.sleep(1)

    # Step 1: Enable ALL keyPopulation checkboxes first
    driver.execute_script("""
        var checkboxes = document.querySelectorAll('input[name="keyPopulation"]');
        for (var i = 0; i < checkboxes.length; i++) {
            checkboxes[i].disabled = false;
            checkboxes[i].readOnly = false;
            checkboxes[i].removeAttribute('disabled');
            checkboxes[i].removeAttribute('readonly');
            // Remove disabled class from parent label
            var label = checkboxes[i].closest('label');
            if (label) {
                label.classList.remove('disabled');
            }
        }
    """)
    time.sleep(0.3)

    # Step 2: If setting something other than "Not Applicable", first uncheck "Not Applicable"
    if value != "Not Applicable":
        driver.execute_script("""
            var checkboxes = document.querySelectorAll('input[name="keyPopulation"]');
            for (var i = 0; i < checkboxes.length; i++) {
                if (checkboxes[i].value === 'Not Applicable' && checkboxes[i].checked) {
                    // Click the label to uncheck
                    var label = checkboxes[i].closest('label');
                    if (label) label.click();
                    break;
                }
            }
        """)
        time.sleep(0.3)

    # Step 3: Click the target checkbox label
    result = driver.execute_script("""
        var value = arguments[0];
        var checkboxes = document.querySelectorAll('input[name="keyPopulation"]');
        var target = null;
        var targetIndex = -1;
        for (var i = 0; i < checkboxes.length; i++) {
            if (checkboxes[i].value === value) {
                target = checkboxes[i];
                targetIndex = i;
                break;
            }
        }
        if (!target) return 'not found: ' + value;

        // Click the label (parent of input)
        var label = target.closest('label');
        if (label) {
            label.click();
            return 'label clicked: ' + target.value + ' checked=' + target.checked;
        }

        // Fallback: direct click on input
        target.click();
        return 'input clicked: ' + target.value + ' checked=' + target.checked;
    """, value)

    log.info(f"  Key Population click result: {result}")

    # Step 4: Verify and retry if not checked
    time.sleep(0.5)
    verified = driver.execute_script("""
        var value = arguments[0];
        var checkboxes = document.querySelectorAll('input[name="keyPopulation"]');
        for (var i = 0; i < checkboxes.length; i++) {
            if (checkboxes[i].value === value) {
                return checkboxes[i].checked;
            }
        }
        return false;
    """, value)

    if verified:
        log.info(f"  OK Key Population -> {value} (verified checked=true)")
        return True

    # Step 5: Retry with direct input click if label click didn't work
    log.info(f"  Retrying Key Population -> {value} with direct click...")
    result2 = driver.execute_script("""
        var value = arguments[0];
        var checkboxes = document.querySelectorAll('input[name="keyPopulation"]');
        for (var i = 0; i < checkboxes.length; i++) {
            if (checkboxes[i].value === value) {
                var cb = checkboxes[i];
                // Ensure enabled
                cb.disabled = false;
                cb.readOnly = false;
                cb.removeAttribute('disabled');
                cb.removeAttribute('readonly');

                // Force checked state
                cb.checked = true;

                // Dispatch events
                cb.dispatchEvent(new Event('input', {bubbles: true}));
                cb.dispatchEvent(new Event('change', {bubbles: true}));

                // Also click the label
                var label = cb.closest('label');
                if (label) {
                    label.classList.remove('disabled');
                    label.click();
                }

                return 'retry: ' + cb.value + ' checked=' + cb.checked;
            }
        }
        return 'not found';
    """, value)
    log.info(f"  Key Population retry result: {result2}")

    # Final verification
    time.sleep(0.5)
    final = driver.execute_script("""
        var value = arguments[0];
        var checkboxes = document.querySelectorAll('input[name="keyPopulation"]');
        for (var i = 0; i < checkboxes.length; i++) {
            if (checkboxes[i].value === value) {
                return checkboxes[i].checked;
            }
        }
        return false;
    """, value)

    if final:
        log.info(f"  OK Key Population -> {value} (final verified)")
        return True

    # XPath fallback
    label_xpaths = [
        f'//input[@name="keyPopulation" and @value="{value}"]/parent::label',
        f'//label[contains(text(),"{value}")]',
    ]
    for xpath in label_xpaths:
        try:
            el = wait_and_find(driver, By.XPATH, xpath, 3)
            driver.execute_script("""
                arguments[0].disabled = false;
                arguments[0].removeAttribute('disabled');
                arguments[0].classList.remove('disabled');
                var input = arguments[0].querySelector('input');
                if (input) {
                    input.disabled = false;
                    input.readOnly = false;
                    input.removeAttribute('disabled');
                    input.removeAttribute('readonly');
                }
            """, el)
            time.sleep(0.3)
            js_click(driver, el)
            time.sleep(0.3)
            # Verify
            is_checked = driver.execute_script("""
                var input = arguments[0].querySelector('input');
                return input ? input.checked : false;
            """, el)
            if is_checked:
                log.info(f"  OK Key Population -> {value} (xpath click)")
                return True
        except Exception:
            continue

    log.warning(f"  SKIP Key Population -> {value}: {result}")
    return False


def fill_health_facility(driver, value):
    """Health Facility fill karo - proper events trigger honge taaki Village load ho."""
    log.info(f"  Setting Health Facility -> {value}")

    for attempt in range(2):
        try:
            # Scroll to Health Facility
            driver.execute_script("""
                var labels = document.querySelectorAll('label.control-label');
                for (var i = 0; i < labels.length; i++) {
                    if (labels[i].textContent.includes('Health Facility')) {
                        labels[i].scrollIntoView({block: 'center'});
                        break;
                    }
                }
            """)
            time.sleep(0.5)

            # Use fill_vselect - proper events trigger honge, Village dropdown load hoga
            return fill_vselect(driver, "Health Facility", value)
        except StaleElementReferenceException:
            log.warning(f"  Health Facility stale element, retrying (attempt {attempt+1})...")
            time.sleep(1)
            dismiss_enrollment_modals(driver)

    log.warning(f"  FAIL Health Facility after {attempt+1} attempts")
    return False


def process_row(driver, record, row_num):
    log.info(f"\n{'='*50}")
    log.info(f"Row {row_num}: {record.get('Name','')}")
    log.info(f"{'='*50}")

    name = str(record.get("Name", "") or "").strip()

    if not name:
        log.warning(f"  SKIP Row {row_num}: Name is empty")
        return False

    time.sleep(1)

    mobile = str(record.get("Mobile", "") or "").strip()
    father = str(record.get("Father/Husband Name", "") or "").strip()
    age = str(record.get("Age", "") or "").strip()
    gender = str(record.get("Gender", "") or "").strip()
    marital = str(record.get("Marital Status", "") or "").strip()
    symptoms_str = str(record.get("Symptoms", "") or "").strip()
    key_pop_str = str(record.get("Key Population", "") or "").strip()
    village = str(record.get("Village", "") or "").strip()

    click_radio_or_checkbox(driver, "I want to add a person from", "Public")
    click_radio_or_checkbox(driver, "Case Finding", "Passive (Routine programme)")
    click_radio_or_checkbox(driver, "Caste", "Other")
    click_radio_or_checkbox(driver, "Area", "Rural")

    parts = name.split(maxsplit=1)
    if parts:
        fill_text(driver, "First Name", parts[0])
        if len(parts) > 1:
            fill_text(driver, "Middle & Last Name", parts[1])

    if father:
        xpath = '//div[contains(@class,"form-group")]//label[contains(@class,"control-label") and contains(text(),"Father")]/following::input[1]'
        try:
            el = wait_and_find(driver, By.XPATH, xpath, 10)
            js_set_val(driver, el, father)
            log.info("  OK Father/Husband's Name")
        except Exception as e:
            log.warning(f"  FAIL Father/Husband's Name: {e}")

    if age:
        fill_text(driver, "Age", age)
    if gender:
        click_radio_or_checkbox(driver, "Gender", gender)
    if mobile:
        fill_text(driver, "Primary Phone", mobile)

    fill_text(driver, "Address", DEFAULT_ADDRESS)
    fill_text(driver, "Pincode", DEFAULT_PINCODE)

    if marital:
        click_radio_or_checkbox(driver, "Marital Status", marital)

    fill_select(driver, "Occupation", "Unknown")
    click_radio_or_checkbox(driver, "Socioeconomic Status", "Unknown")

    if symptoms_str:
        for s in [x.strip() for x in symptoms_str.split(",") if x.strip()]:
            handle_symptom(driver, "Symptoms", s)

    if key_pop_str:
        key_pop_filled = False
        for k in [x.strip() for x in key_pop_str.split(",") if x.strip()]:
            mapped = k
            for search, replace in KEY_POP_MAP.items():
                if search.lower() in k.lower() or k.lower() in search.lower():
                    mapped = replace
                    break
            if click_key_population(driver, mapped):
                key_pop_filled = True

        if not key_pop_filled:
            log.warning(f"  Key Population '{key_pop_str}' not found, using default: Other")
            click_key_population(driver, "Other")
    else:
        click_key_population(driver, "Not Applicable")

    dismiss_enrollment_modals(driver)

    # Health Facility fill karo
    fill_health_facility(driver, DEFAULT_HEALTH_FACILITY)

    # Village dropdown API se load hota hai - directly Bermi try karo wait ke saath
    village_filled = False
    village_name = village if village else "Bermi"

    for attempt in range(3):
        if fill_vselect(driver, "Village", village_name):
            village_filled = True
            break
        log.warning(f"  Village '{village_name}' attempt {attempt+1} failed, retrying...")
        time.sleep(1)

    # Agar original village nahi mila toh default Bermi try karo
    if not village_filled and village and village.lower() != "bermi":
        log.warning(f"  Village '{village}' not found, trying default: Bermi")
        if fill_vselect(driver, "Village", "Bermi"):
            village_filled = True

    if not village_filled:
        log.warning("  Village could not be filled")

    submit_form(driver)
    time.sleep(0.5)
    cancel_abha_popup(driver)
    go_to_enrollment(driver)


def find_excel():
    files = glob.glob(os.path.join(DATA_FOLDER, "*.xlsx"))
    if not files:
        log.error(f"No Excel file found in {DATA_FOLDER}")
        return None, None
    fp = files[0]
    wb = load_workbook(fp, data_only=True)
    sheet = wb.sheetnames[0]
    log.info(f"Using: {os.path.basename(fp)} -> Sheet: {sheet}")
    return fp, sheet


def read_excel(filepath, sheet_name):
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    headers = [cell.value for cell in ws[1]]
    data = []
    for row in rows:
        if not row[0]:
            continue
        data.append(dict(zip(headers, row)))
    log.info(f"Loaded {len(data)} rows")
    return data, headers


def update_excel_after_success(filepath, sheet_name, successful_names):
    """Original Excel se successful rows delete karo."""
    if not successful_names:
        return
    try:
        wb = load_workbook(filepath)
        ws = wb[sheet_name]

        # Row numbers (1-indexed, header is row 1)
        rows_to_delete = []
        for row_idx in range(2, ws.max_row + 1):
            name = ws.cell(row=row_idx, column=1).value
            if name and str(name).strip() in successful_names:
                rows_to_delete.append(row_idx)

        # Delete from bottom to top (so indices don't shift)
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx, 1)

        wb.save(filepath)
        log.info(f"  Excel updated: {len(rows_to_delete)} rows deleted")
    except Exception as e:
        log.error(f"  Excel update failed: {e}")


def save_failed_records(failed_records, headers):
    """Failed records ko failed_records.xlsx mein save karo."""
    if not failed_records:
        return
    try:
        fp = os.path.join(DATA_FOLDER, "failed_records.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Failed Records"

        # Headers with error reason
        ws.append(headers + ["Error"])

        # Add failed records
        for rec in failed_records:
            row = [rec.get(h, "") for h in headers]
            row.append(rec.get("_error", ""))
            ws.append(row)

        wb.save(fp)
        log.info(f"  Failed records saved to: {fp}")
    except Exception as e:
        log.error(f"  Save failed records error: {e}")


def main():
    log.info("=== Nikshay Data Entry Bot ===")

    if not NIKSHAY_USERNAME or not NIKSHAY_PASSWORD:
        log.error("NIKSHAY_USERNAME or NIKSHAY_PASSWORD not set! Check config.env")
        return

    fp, sheet = find_excel()
    if not fp:
        return
    records, headers = read_excel(fp, sheet)
    if not records:
        log.error("No records!")
        return

    total = len(records)
    successful_names = set()
    failed_records = []

    log.info(f"Total records to process: {total}")

    driver = init_driver()
    try:
        if not login(driver):
            log.error("Login failed! Check debug screenshots and debug_page_source.html")
            take_screenshot(driver, "FINAL_LOGIN_FAIL")
            if not HEADLESS:
                input("Press Enter to close browser...")
            return
        go_to_enrollment(driver)

        for i, rec in enumerate(records[START_ROW - 1:], START_ROW):
            retries = 0
            success = False
            while retries < MAX_RETRIES and not success:
                try:
                    result = process_row(driver, rec, i)
                    if result is False:
                        log.info(f"  Row {i} skipped (empty)")
                        success = True
                    else:
                        success = True
                        successful_names.add(str(rec.get("Name", "")).strip())
                        filled = len(successful_names)
                        log.info(f"  PROGRESS: {filled}/{total} filled ({total - filled} remaining)")
                except Exception as e:
                    retries += 1
                    log.error(f"Row {i} failed (attempt {retries}): {e}")
                    traceback.print_exc()
                    try:
                        dismiss_popups(driver)
                    except Exception:
                        pass
                    if retries < MAX_RETRIES:
                        log.info("  Retrying in 10s...")
                        time.sleep(10)
                        go_to_enrollment(driver)
                    else:
                        log.error(f"  Row {i} skipped after {MAX_RETRIES} attempts")
                        rec["_error"] = str(e)
                        failed_records.append(rec)

        # === Summary ===
        log.info("\n" + "=" * 50)
        log.info("=== SUMMARY ===")
        log.info(f"  Total:     {total}")
        log.info(f"  Filled:    {len(successful_names)}")
        log.info(f"  Failed:    {len(failed_records)}")
        log.info("=" * 50)

        # Update Excel - delete successful rows
        if successful_names:
            log.info("Updating original Excel (removing filled records)...")
            update_excel_after_success(fp, sheet, successful_names)

        # Save failed records
        if failed_records:
            log.info("Saving failed records...")
            save_failed_records(failed_records, headers)

        log.info("\n=== All done ===")
    finally:
        if not HEADLESS:
            input("Press Enter to close browser...")
        driver.quit()


if __name__ == "__main__":
    main()
